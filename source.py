import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

#name=[]
# program read information from people.csv file and put all data in name list.

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)




found_element = driver.find_element(By.ID, "input")
found_element.send_keys("Adrienne Lambert")
time.sleep(2)
search_element=driver.find_element(By.CLASS_NAME, "btn")
search_element.click()
time.sleep(3)

result_element=driver.find_element(By.ID, "output" )
time.sleep(5)
result = result_element.get_attribute("value")
print(result)
from openpyxl import load_workbook
wb=load_workbook('salary.xlsx')
ws=wb.active
rindas=ws.max_row
saraksts=[]
num=0
n=0
for r in range(2,rindas+1):
    a=(ws['a'+str(r)].value)
    b=(ws['b'+str(r)].value)
    
    if a==result:
        saraksts.append(int(b))
        num=sum(saraksts)
print(num)
#nolasīt xlsx ar openpyxl
#ideālais variants ir:
#1 eksemplārs ar originālajiem datiem
#2 eksemplārs ar rezultātiem

    #vārds uzvārds
    #get lapu
    #iestatījumu maiņa
    #ar XPATH atrast option un click
    #ievadīt vārdu uzvārdu id="input"
    #nolasīt rezultātu iekšā id="input"
    #aizvietojama 2.datu kopijā attiecīgus vārdus
#masīvu 2.datu kopijai var pāarvērst par data frame (pandas biblioteka)
#kopējā alga katram darbiniekam
